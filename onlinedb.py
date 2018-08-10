#onlinedb.py populatedb C:\Users\DELL\Desktop\students.xlsx C:\Users\DELL\Desktop\marks.xlsx



from openpyxl import load_workbook
import click
import os
import django
import MySQLdb


host='localhost'
username = 'pranadeep'
password = 'pranadeep'


os.environ.setdefault("DJANGO_SETTINGS_MODULE","onlineproject.settings")
django.setup()


from onlineapp.models import *

def findDBName(str):
    str=str.replace("ol2016_",'')
    str=str.replace("_mock",'')
    return str[str.find('_')+1:]


@click.group()
def cli():
    pass


@cli.command()
def createdb():
    db = MySQLdb.connect(host, username, password)
    cursor = db.cursor()
    sql = 'create database temp_db'
    cursor.execute(sql)
    db.commit()
    db.close()
    os.system("python manage.py makemigrations")
    os.system("python manage.py migrate")


@cli.command()
def dropdb():
    db = MySQLdb.connect(host, username, password)
    cursor = db.cursor()
    sql = 'drop database if exists temp_db'
    cursor.execute(sql)
    db.commit()
    db.close()

@cli.command()
def cleardata():
    College.objects.all().delete()


@cli.command()
@click.argument('f1',nargs=1)
@click.argument('f2',nargs=1)
def populatedb(f1,f2):
    studentsWorkbook = load_workbook(f1)
    CollegeWorksheet = studentsWorkbook.get_sheet_by_name('Colleges')
    for i in range(1, CollegeWorksheet.max_row):
        collegeData = []
        for j in range(0, CollegeWorksheet.max_column):
            cell = CollegeWorksheet.cell(row=i + 1, column=j + 1)
            collegeData.append(cell.value)
        c = College(name=collegeData[0], acronym=collegeData[1], location=collegeData[2], contact=collegeData[3])
        c.save()


    studentWorksheet = studentsWorkbook.get_sheet_by_name('Current')
    for i in range(1, studentWorksheet.max_row):
        studentData = []
        for j in range(0, studentWorksheet.max_column):
            cell = studentWorksheet.cell(row=i + 1, column=j + 1)
            studentData.append(cell.value)
        s = Student(name=studentData[0],dropped_out = False, college=College.objects.get(acronym = studentData[1]), db_folder=studentData[3], email=studentData[2])
        s.save()

    studentWorksheet = studentsWorkbook.get_sheet_by_name('Deletions')
    for i in range(1, studentWorksheet.max_row):
        studentData = []
        for j in range(0, studentWorksheet.max_column):
            cell = studentWorksheet.cell(row=i + 1, column=j + 1)
            studentData.append(cell.value)
        s = Student(name=studentData[0], dropped_out=True, college=College.objects.get(acronym=studentData[1]),
                    db_folder=studentData[3], email=studentData[2])
        s.save()


    markssWorkbook = load_workbook(f2)
    marksWorksheet = markssWorkbook.get_sheet_by_name('dumpedsheet')
    for i in range(1, marksWorksheet.max_row):
        marksData = []
        for j in range(0, marksWorksheet.max_column):
            cell = marksWorksheet.cell(row=i + 1, column=j + 1)
            if j == 0:
                marksData.append(cell.value)
            else:
                marksData.append(int(cell.value))
        marksData[0] = findDBName(marksData[0])
        m = MockTest1(student=Student.objects.get(db_folder = marksData[0]), problem1=marksData[1], problem2=marksData[2],
                      problem3=marksData[3], problem4=marksData[4], total=marksData[5])
        m.save()

if __name__=='__main__':
    cli()